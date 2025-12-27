from __future__ import annotations

import os
import pathlib
from contextlib import nullcontext as does_not_raise
from pathlib import Path

import openpyxl
import pytest
from femto.curves import sin
from femto.device import Cell
from femto.device import Device
from femto.helpers import delete_folder
from femto.helpers import flatten
from femto.marker import Marker
from femto.trench import Trench
from femto.trench import TrenchColumn
from femto.waveguide import NasuWaveguide
from femto.waveguide import Waveguide
from shapely.geometry import Polygon


@pytest.fixture
def ss_param() -> dict:
    return dict(
        book_name='custom_book_name.xlsx',
        sheet_name='custom_sheet_name',
        columns_names=['name', 'power', 'speed', 'scan', 'depth', 'int_dist', 'yin', 'yout', 'obs'],
    )


@pytest.fixture
def gc_param() -> dict:
    p = dict(
        filename='testCell.pgm',
        laser='PHAROS',
        shift_origin=(0.5, 0.5),
        samplesize=(25, 1),
    )
    return p


@pytest.fixture
def list_wg() -> list[Waveguide]:
    PARAM_WG = dict(speed=20, radius=25, pitch=0.080, int_dist=0.007, samplesize=(25, 3))

    coup = [Waveguide(**PARAM_WG) for _ in range(5)]
    for i, wg in enumerate(coup):
        wg.start([-2, i * wg.pitch, 0.035])
        wg.linear([5, 0, 0])
        wg.coupler(dy=(-1) ** i * wg.dy_bend, dz=0, fx=sin)
        wg.coupler(dy=(-1) ** i * wg.dy_bend, dz=0, fx=sin)
        wg.linear([5, 0, 0])
        wg.end()
    return coup


@pytest.fixture
def list_mk() -> list[Marker]:
    PARAM_MK = dict(scan=1, speed=2, speed_pos=5, speed_closed=5, depth=0.000, lx=1, ly=1)
    markers = []
    for x, y in zip(range(4, 8), range(3, 7)):
        m = Marker(**PARAM_MK)
        m.cross([x, y])
        markers.append(m)
    return markers


@pytest.fixture
def list_tcol(list_wg) -> list[TrenchColumn]:
    PARAM_TC = dict(
        length=1.0,
        base_folder='',
        y_min=-0.1,
        y_max=4 * 0.08 + 0.1,
        u=[30.339, 32.825],
    )
    x_c = [3, 7.5, 10.5]
    t_col = []
    for x in x_c:
        T = TrenchColumn(x_center=x, **PARAM_TC)
        T.dig_from_waveguide(flatten([list_wg]))
        t_col.append(T)
    return t_col


@pytest.fixture
def device(gc_param) -> Device:
    return Device(**gc_param)


@pytest.fixture
def cell() -> Cell:
    return Cell()


# Cell
def test_cell_init(cell) -> None:
    assert cell.name == 'base'
    assert cell.description is None
    assert isinstance(cell._objs, dict)
    assert cell._objs == {
        TrenchColumn: [],
        Marker: [],
        Waveguide: [],
        NasuWaveguide: [],
    }


@pytest.mark.parametrize(
    'name, exp',
    [
        ('base', 'base'),
        ('base test', 'base-test'),
        ('TEST-TEST', 'test-test'),
        ('test_test', 'test_test'),
        ('A B C', 'a-b-c'),
        ('ab C', 'ab-c'),
        ('ab C  ', 'ab-c'),
        ('ab C     ', 'ab-c'),
        ('ab C     d', 'ab-c-----d'),
    ],
)
def test_cell_rename(name, exp) -> None:
    cell = Cell(name=name)
    assert cell.name == exp


@pytest.mark.parametrize(
    'name, exp',
    [
        ('base', 'base'),
        ('base test', 'base-test'),
        ('TEST-TEST', 'test-test'),
        ('test_test', 'test_test'),
        ('A B C', 'a-b-c'),
        ('ab C', 'ab-c'),
    ],
)
def test_cell_repr(name, exp) -> None:
    cell = Cell(name=name)
    assert cell.__repr__() == f'Cell {exp}'


def test_cell_objects_empty(cell) -> None:
    assert isinstance(cell.objects, dict)
    assert cell.objects == {
        TrenchColumn: [],
        Marker: [],
        Waveguide: [],
        NasuWaveguide: [],
    }


def test_cell_objects(cell, list_wg, list_mk) -> None:
    cell.add(list_wg)
    cell.add(list_mk)
    cell.add([list_wg, list_mk])
    obj = cell.objects
    assert obj[TrenchColumn] == []
    assert obj[NasuWaveguide] == []
    assert obj[Waveguide] == flatten([list_wg, list_wg])
    assert obj[Marker] == flatten([list_mk, list_mk])


@pytest.mark.parametrize(
    'inp, expectation',
    [
        ([Waveguide(), Waveguide()], does_not_raise()),
        ([Marker(), [Waveguide(), Waveguide()]], does_not_raise()),
        ([Trench(Polygon()), Waveguide()], pytest.raises(TypeError)),
        ([], does_not_raise()),
    ],
)
def test_cell_parse_objects_raise(cell, inp, expectation) -> None:
    with expectation:
        cell.parse_objects(inp)


def test_cell_add(list_wg, list_mk, list_tcol) -> None:
    cell = Cell()
    cell.add(list_wg)
    assert cell.objects[Waveguide] == list_wg
    assert cell.objects[TrenchColumn] == []
    assert cell.objects[Marker] == []
    assert cell.objects[TrenchColumn] == []
    assert cell.objects[NasuWaveguide] == []
    del cell

    cell = Cell()
    cell.add(list_wg)
    assert cell.objects[Waveguide] == list_wg
    assert cell.objects[TrenchColumn] == []
    assert cell.objects[Marker] == []
    assert cell.objects[NasuWaveguide] == []

    cell.add(list_mk)
    assert cell.objects[Waveguide] == list_wg
    assert cell.objects[TrenchColumn] == []
    assert cell.objects[Marker] == list_mk
    assert cell.objects[NasuWaveguide] == []

    cell.add(list_tcol)
    assert cell.objects[Waveguide] == list_wg
    assert cell.objects[TrenchColumn] == list_tcol
    assert cell.objects[Marker] == list_mk
    assert cell.objects[NasuWaveguide] == []
    del cell

    cell = Cell()
    cell.add(list_wg)
    cell.add(list_mk)
    cell.add(list_tcol)
    assert cell.objects[Waveguide] == list_wg
    assert cell.objects[TrenchColumn] == list_tcol
    assert cell.objects[Marker] == list_mk
    assert cell.objects[NasuWaveguide] == []
    del cell

    cell = Cell()
    cell.add(list_wg)
    cell.add(list_mk)
    cell.add(list_wg)
    cell.add(list_tcol)
    cell.add(list_tcol)
    cell.add(list_wg)
    assert cell.objects[Waveguide] == list_wg * 3
    assert cell.objects[TrenchColumn] == list_tcol * 2
    assert cell.objects[Marker] == list_mk
    assert cell.objects[NasuWaveguide] == []
    del cell


@pytest.mark.parametrize(
    'inp, expectation',
    [
        ([Waveguide(), Waveguide()], does_not_raise()),
        ([Marker(), [Waveguide(), Waveguide()]], does_not_raise()),
        ([Trench(Polygon()), Waveguide()], pytest.raises(TypeError)),
        ([], does_not_raise()),
        ([1, 2, Waveguide(), Marker], pytest.raises(ValueError)),
    ],
)
def test_cell_ass_raise(cell, inp, expectation) -> None:
    with expectation:
        cell.add(inp)


# Device
def test_device_init(device, gc_param) -> None:
    assert device.cells_collection == {}
    assert device.fig is None
    assert device.fabrication_time == 0.0

    assert device._param == gc_param
    assert device._print_angle_warning is True
    assert device._print_base_cell_warning is True


def test_device_from_dict(device, gc_param) -> None:
    dev = Device.from_dict(gc_param)
    assert dev.cells_collection == device.cells_collection
    assert dev._param == device._param
    assert dev._param == gc_param
    assert dev.fig is device.fig
    assert dev.fabrication_time == dev.fabrication_time


def test_device_from_dict_w_kwargs(device, gc_param) -> None:
    dev = Device.from_dict(gc_param, filename='test_kwargs', flip_x=True)

    param = device._param
    param.update({'filename': 'test_kwargs', 'flip_x': True})

    assert dev.cells_collection == device.cells_collection
    assert dev._param == param
    assert dev.fig is device.fig
    assert dev.fabrication_time == dev.fabrication_time


def test_device_keys(device) -> None:
    c1 = Cell(name='name1')
    c2 = Cell(name='name2')
    c3 = Cell(name='name3')
    c4 = Cell(name='name4')

    device.add([c1, c2, c3, c4])
    assert device.keys == ['name1', 'name2', 'name3', 'name4']


def test_device_remove_cell(device, cell) -> None:
    c2 = Cell(name='c2')
    c3 = Cell(name='c3')
    device.add_cell(cell)
    device.add_cell(c2)
    device.add_cell(c3)

    device.remove_cell(c2)
    assert 'c2' not in device.cells_collection.keys()

    device.remove_cell(c3)
    assert 'c3' not in device.cells_collection.keys()

    device.remove_cell(cell)
    assert cell.name not in device.cells_collection.keys()
    assert list(device.cells_collection.keys()) == []


def test_device_remove_cell_not_present(device, cell) -> None:
    c2 = Cell(name='c2')
    c3 = Cell(name='c3')
    device.add_cell(c2)
    device.add_cell(c3)

    with pytest.raises(KeyError):
        device.remove_cell(cell)


def test_device_add_cell_same_name(device, cell) -> None:
    c2 = cell
    with pytest.raises(KeyError):
        device.add_cell(cell)
        device.add_cell(c2)


def test_device_add_cell(device, cell, list_tcol, list_wg) -> None:
    c2 = Cell(name='new')
    c2.add(list_wg)
    device.add(list_tcol)

    device.add_cell(c2)
    assert len(device.cells_collection.values()) == 2
    assert all(keys in ['base', 'new'] for keys in device.cells_collection.keys())

    assert device.cells_collection['base'].objects[Waveguide] == []
    assert device.cells_collection['base'].objects[NasuWaveguide] == []
    assert device.cells_collection['base'].objects[Marker] == []
    assert device.cells_collection['base'].objects[TrenchColumn] == list_tcol

    assert device.cells_collection['new'].objects[Waveguide] == list_wg
    assert device.cells_collection['new'].objects[NasuWaveguide] == []
    assert device.cells_collection['new'].objects[Marker] == []
    assert device.cells_collection['new'].objects[TrenchColumn] == []


def test_device_add_to_cell(device, cell, list_wg, list_mk) -> None:
    c2 = Cell(name='new')
    c2.add(list_wg)
    device.add(c2)

    assert device.cells_collection['new'].objects[Waveguide] == list_wg
    assert device.cells_collection['new'].objects[NasuWaveguide] == []
    assert device.cells_collection['new'].objects[Marker] == []
    assert device.cells_collection['new'].objects[TrenchColumn] == []

    device.add_to_cell(key=c2.name, obj=list_mk)
    assert len(device.cells_collection.values()) == 1
    assert list(device.cells_collection.keys()) == ['new']

    assert device.cells_collection['new'].objects[Waveguide] == list_wg
    assert device.cells_collection['new'].objects[NasuWaveguide] == []
    assert device.cells_collection['new'].objects[Marker] == list_mk
    assert device.cells_collection['new'].objects[TrenchColumn] == []


def test_device_add_to_non_existing_cell(device, cell, list_wg, list_mk) -> None:
    c2 = Cell(name='new')
    c2.add(list_wg)
    device.add(c2)

    assert device.cells_collection['new'].objects[Waveguide] == list_wg
    assert device.cells_collection['new'].objects[NasuWaveguide] == []
    assert device.cells_collection['new'].objects[Marker] == []
    assert device.cells_collection['new'].objects[TrenchColumn] == []

    device.add_to_cell(key='new2', obj=list_mk)
    assert len(device.cells_collection.values()) == 2
    assert list(device.cells_collection.keys()) == ['new', 'new2']

    assert device.cells_collection['new2'].objects[Waveguide] == []
    assert device.cells_collection['new2'].objects[NasuWaveguide] == []
    assert device.cells_collection['new2'].objects[Marker] == list_mk
    assert device.cells_collection['new2'].objects[TrenchColumn] == []


def test_device_add_list_femtobjs(device, list_wg, list_mk, list_tcol) -> None:
    listone = flatten([list_wg, list_mk, list_tcol])
    device.add(listone)

    assert device.cells_collection['base'].objects[Waveguide] == list_wg
    assert device.cells_collection['base'].objects[NasuWaveguide] == []
    assert device.cells_collection['base'].objects[Marker] == list_mk
    assert device.cells_collection['base'].objects[TrenchColumn] == list_tcol


def test_device_add_list_femtobjs_multi(device, list_wg, list_mk, list_tcol) -> None:
    listone = flatten([list_wg, list_mk, list_tcol, list_wg, list_wg])
    device.add(listone)

    assert device.cells_collection['base'].objects[Waveguide] == list_wg * 3
    assert device.cells_collection['base'].objects[NasuWaveguide] == []
    assert device.cells_collection['base'].objects[Marker] == list_mk
    assert device.cells_collection['base'].objects[TrenchColumn] == list_tcol


def test_device_add_single_femtobjs(device, list_wg, list_mk, list_tcol) -> None:
    wg, mk, tc = list_wg[0], list_mk[0], list_tcol[0]
    device.add(wg)
    device.add(mk)
    device.add(tc)

    assert device.cells_collection['base'].objects[Waveguide] == [wg]
    assert device.cells_collection['base'].objects[NasuWaveguide] == []
    assert device.cells_collection['base'].objects[Marker] == [mk]
    assert device.cells_collection['base'].objects[TrenchColumn] == [tc]


def test_device_add_single_cell(device) -> None:
    cell = Cell(name='test')
    device.add(cell)

    assert len(device.cells_collection.values()) == 1
    assert list(device.cells_collection.keys()) == ['test']
    assert device.cells_collection['test'].objects[Waveguide] == []
    assert device.cells_collection['test'].objects[NasuWaveguide] == []
    assert device.cells_collection['test'].objects[Marker] == []
    assert device.cells_collection['test'].objects[TrenchColumn] == []


def test_device_add_single_cell_multi(device) -> None:
    cell = Cell(name='test')
    device.add(cell)
    cell = Cell(name='test1')
    device.add(cell)
    cell = Cell(name='test2')
    device.add(cell)

    assert len(device.cells_collection.values()) == 3
    assert list(device.cells_collection.keys()) == ['test', 'test1', 'test2']
    assert device.cells_collection['test'].objects[Waveguide] == []
    assert device.cells_collection['test'].objects[NasuWaveguide] == []
    assert device.cells_collection['test'].objects[Marker] == []
    assert device.cells_collection['test'].objects[TrenchColumn] == []

    assert device.cells_collection['test1'].objects[Waveguide] == []
    assert device.cells_collection['test1'].objects[NasuWaveguide] == []
    assert device.cells_collection['test1'].objects[Marker] == []
    assert device.cells_collection['test1'].objects[TrenchColumn] == []

    assert device.cells_collection['test2'].objects[Waveguide] == []
    assert device.cells_collection['test2'].objects[NasuWaveguide] == []
    assert device.cells_collection['test2'].objects[Marker] == []
    assert device.cells_collection['test2'].objects[TrenchColumn] == []


def test_device_add_list_cell_multi(device) -> None:
    cell1 = Cell(name='test')
    cell2 = Cell(name='test1')
    cell3 = Cell(name='test2')
    device.add([cell1, cell2, cell3])

    assert len(device.cells_collection.values()) == 3
    assert list(device.cells_collection.keys()) == ['test', 'test1', 'test2']
    assert device.cells_collection['test'].objects[Waveguide] == []
    assert device.cells_collection['test'].objects[NasuWaveguide] == []
    assert device.cells_collection['test'].objects[Marker] == []
    assert device.cells_collection['test'].objects[TrenchColumn] == []

    assert device.cells_collection['test1'].objects[Waveguide] == []
    assert device.cells_collection['test1'].objects[NasuWaveguide] == []
    assert device.cells_collection['test1'].objects[Marker] == []
    assert device.cells_collection['test1'].objects[TrenchColumn] == []

    assert device.cells_collection['test2'].objects[Waveguide] == []
    assert device.cells_collection['test2'].objects[NasuWaveguide] == []
    assert device.cells_collection['test2'].objects[Marker] == []
    assert device.cells_collection['test2'].objects[TrenchColumn] == []


@pytest.mark.parametrize(
    'elem, exp',
    [
        ([Waveguide(), 1, 2, 3], pytest.raises(TypeError)),
        ([Waveguide(), Marker(), Marker()], does_not_raise()),
        ([[[Waveguide(), Marker(), Marker()]]], does_not_raise()),
        (NasuWaveguide(), does_not_raise()),
        ([NasuWaveguide(), Waveguide()], does_not_raise()),
        ([Waveguide()], does_not_raise()),
        ([None], pytest.raises(TypeError)),
        (None, pytest.raises(TypeError)),
    ],
)
def test_device_add_raise(device, elem, exp) -> None:
    with exp:
        device.add(elem)


def test_plot2d_save(device, list_wg, list_mk, list_tcol) -> None:
    device.add([list_wg, list_mk, list_tcol])

    device.plot2d(save=False, show=False)
    assert not (Path('.').cwd() / 'scheme.html').is_file()
    assert device.fig is not None


def test_plot2d_save_true(device, list_wg, list_mk, list_tcol) -> None:
    device.add([list_wg, list_mk, list_tcol])

    device.plot2d(show=False, save=True)
    assert (Path('.').cwd() / 'scheme.html').is_file()
    assert device.fig is not None
    (Path('.').cwd() / 'scheme.html').unlink()


def test_plot3d_save(device, list_wg, list_mk, list_tcol) -> None:
    device.add([list_wg, list_mk, list_tcol])

    device.plot3d(save=False, show=False)
    assert not (Path('.').cwd() / 'scheme.html').is_file()
    assert device.fig is not None


def test_plot3d_save_true(device, list_wg, list_mk, list_tcol) -> None:
    device.add([list_wg, list_mk, list_tcol])

    device.plot3d(show=False, save=True)
    assert (Path('.').cwd() / 'scheme.html').is_file()
    assert device.fig is not None
    (Path('.').cwd() / 'scheme.html').unlink()


def test_device_pgm(device, list_wg, list_mk) -> None:
    device.add([list_wg, list_mk])
    device.pgm()
    assert (Path().cwd() / 'testCell_WG.pgm').is_file()
    assert (Path().cwd() / 'testCell_MK.pgm').is_file()
    (Path().cwd() / 'testCell_WG.pgm').unlink()
    (Path().cwd() / 'testCell_MK.pgm').unlink()


def test_device_pgm_verbose(device, list_wg, list_mk) -> None:
    device.add([list_wg, list_mk])
    device.pgm(verbose=True)
    assert (Path().cwd() / 'testCell_WG.pgm').is_file()
    assert (Path().cwd() / 'testCell_MK.pgm').is_file()
    (Path().cwd() / 'testCell_WG.pgm').unlink()
    (Path().cwd() / 'testCell_MK.pgm').unlink()


def test_device_pgm_custom_folder(device, list_wg, list_mk) -> None:
    cell = Cell(name='test')
    cell.add([list_wg, list_mk])
    device.add(cell)
    device.pgm()
    assert (Path().cwd() / 'TEST_WG.pgm').is_file()
    assert (Path().cwd() / 'TEST_MK.pgm').is_file()
    (Path().cwd() / 'TEST_WG.pgm').unlink()
    (Path().cwd() / 'TEST_MK.pgm').unlink()


def test_device_pgm_custom_folder_and_base(device, list_wg, list_mk) -> None:
    cell = Cell(name='test')
    cell.add([list_wg, list_mk])
    device.add(cell)
    device.add([list_wg, list_mk])
    device.pgm()

    assert (Path().cwd() / 'TEST_WG.pgm').is_file()
    assert (Path().cwd() / 'TEST_MK.pgm').is_file()
    (Path().cwd() / 'TEST_WG.pgm').unlink()
    (Path().cwd() / 'TEST_MK.pgm').unlink()

    assert (Path().cwd() / 'BASE_WG.pgm').is_file()
    assert (Path().cwd() / 'BASE_MK.pgm').is_file()
    (Path().cwd() / 'BASE_WG.pgm').unlink()
    (Path().cwd() / 'BASE_MK.pgm').unlink()


def test_device_xlsx(device, list_wg, list_mk, ss_param) -> None:
    device.add([list_wg, list_mk])
    device.xlsx(**ss_param)
    assert (Path().cwd() / 'custom_book_name.xlsx').is_file()
    (Path().cwd() / 'custom_book_name.xlsx').unlink()


def test_device_xlsx_meta(device, list_wg, list_mk, ss_param) -> None:
    device.add([list_wg, list_mk])
    ss_param['metadata'] = {'laser': 'UWE'}
    device.xlsx(**ss_param)
    assert (Path().cwd() / 'custom_book_name.xlsx').is_file()
    wb = openpyxl.load_workbook(Path().cwd() / 'custom_book_name.xlsx')
    assert wb['custom_sheet_name'].cell(row=25, column=3).value == 'UWE'
    (Path().cwd() / 'custom_book_name.xlsx').unlink()


def test_device_xlsx_meta_ext(device, list_wg, list_mk, ss_param) -> None:
    device.add([list_wg, list_mk])
    ss_param['metadata'] = {'laser': 'UWE'}
    device.xlsx(**ss_param)
    assert (Path().cwd() / 'custom_book_name.xlsx').is_file()
    wb = openpyxl.load_workbook(Path().cwd() / 'custom_book_name.xlsx')
    assert wb['custom_sheet_name'].cell(row=25, column=3).value == 'UWE'
    (Path().cwd() / 'custom_book_name.xlsx').unlink()


def test_device_save_empty(device) -> None:
    assert device.save() is None
    assert not (Path('.').cwd() / 'scheme.html').is_file()


@pytest.mark.parametrize(
    'fn, expected',
    [
        ('test.html', 'test.html'),
        ('TEST.HTML', 'TEST.html'),
        ('test', 'test.html'),
        ('test.svg', 'test.svg'),
        ('test.pdf', 'test.pdf'),
        ('test.jpeg', 'test.jpeg'),
        ('test.png', 'test.png'),
    ],
)
def test_device_save(device, fn, expected) -> None:
    device.plot2d(show=False)
    opt = {'width': 480, 'height': 320, 'scale': 2, 'engine': 'kaleido'}
    assert device.save(fn, opt) is None
    assert (Path('.').cwd() / expected).is_file()
    (Path('.').cwd() / expected).unlink()


def test_device_export(device, list_wg, list_mk) -> None:
    device.add([list_wg, list_mk])
    device.export()
    for i, _ in enumerate(list_wg):
        fn = Path().cwd() / 'EXPORT' / 'BASE' / f'WG_{i + 1:02}.pickle'
        assert fn.is_file()
    for i, _ in enumerate(list_mk):
        fn = Path().cwd() / 'EXPORT' / 'BASE' / f'MK_{i + 1:02}.pickle'
        assert fn.is_file()

    delete_folder(Path().cwd() / 'EXPORT')


def test_device_export_verbose(device, list_wg, list_mk) -> None:
    device.add([list_wg, list_mk])
    device.export()

    for i, _ in enumerate(list_wg):
        fn = Path().cwd() / 'EXPORT' / 'BASE' / f'WG_{i + 1:02}.pickle'
        assert fn.is_file()
    for i, _ in enumerate(list_mk):
        fn = Path().cwd() / 'EXPORT' / 'BASE' / f'MK_{i + 1:02}.pickle'
        assert fn.is_file()

    delete_folder(Path().cwd() / 'EXPORT')


def test_device_load_verbose(device, list_wg, list_mk, gc_param) -> None:
    device.add([list_wg, list_mk])
    device.export()

    fn = Path().cwd() / 'EXPORT' / 'BASE'

    d2 = Device.load_objects(fn, gc_param, verbose=True)
    assert d2.cells_collection['base'].objects[Waveguide]
    assert d2.cells_collection['base'].objects[Marker]
    assert not d2.cells_collection['base'].objects[TrenchColumn]
    assert not d2.cells_collection['base'].objects[NasuWaveguide]

    delete_folder(fn.parent)


def test_device_load_empty(list_wg, list_mk, list_tcol, gc_param) -> None:
    device = Device(**gc_param)
    device.add([list_tcol, list_wg, list_mk])
    device.export()

    # load on a different, empty path
    fn = Path().cwd() / 'EXPORT' / 'BAASE'

    fn.mkdir(parents=True, exist_ok=True)
    (fn / 'cell1').mkdir(parents=True, exist_ok=True)
    (fn / 'cell2').mkdir(parents=True, exist_ok=True)
    (fn / 'cell3').mkdir(parents=True, exist_ok=True)

    d2 = Device.load_objects(fn, gc_param, verbose=True)
    assert not d2.cells_collection['base'].objects[Waveguide]
    assert not d2.cells_collection['base'].objects[Marker]
    assert not d2.cells_collection['base'].objects[TrenchColumn]
    assert not d2.cells_collection['base'].objects[NasuWaveguide]

    delete_folder(fn.parent)


def test_device_export_non_base_cell(device, list_wg, list_mk) -> None:
    cell = Cell(name='test')
    cell.add([list_wg, list_mk])
    device.add(cell)
    device.export()

    for i, _ in enumerate(list_wg):
        fn = Path().cwd() / 'EXPORT' / 'TEST' / f'WG_{i + 1:02}.pickle'
        assert fn.is_file()
    for i, _ in enumerate(list_mk):
        fn = Path().cwd() / 'EXPORT' / 'TEST' / f'MK_{i + 1:02}.pickle'
        assert fn.is_file()

    delete_folder(Path().cwd() / 'EXPORT')


def test_device_load_non_base_cell(device, gc_param, list_wg, list_mk, list_tcol) -> None:
    cell = Cell(name='test')
    cell.add([list_wg, list_mk, list_tcol])
    device.add(cell)
    device.export()
    del device

    fn = Path().cwd() / 'EXPORT' / 'TEST'

    d2 = Device.load_objects(fn, gc_param, verbose=True)
    assert d2.cells_collection['test'].objects[Waveguide]
    assert d2.cells_collection['test'].objects[Marker]
    assert d2.cells_collection['test'].objects[TrenchColumn]
    assert not d2.cells_collection['test'].objects[NasuWaveguide]
    # assert d2.writers[Waveguide].objs == list_wg
    # assert d2.writers[TrenchColumn].objs == list_tcol
    # assert d2.writers[Marker].objs == list_mk

    for root, dirs, files in os.walk(fn):
        for file in files:
            (pathlib.Path(root) / file).unlink()

    delete_folder(Path().cwd() / 'EXPORT')
