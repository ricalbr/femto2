from __future__ import annotations

from contextlib import nullcontext as does_not_raise
from pathlib import Path

import openpyxl
import pytest
from femto import __file__ as fpath
from femto.curves import sin
from femto.marker import Marker
from femto.spreadsheet import Spreadsheet
from femto.waveguide import Waveguide

src_path = Path(fpath).parent
dot_path = Path('.').cwd()


@pytest.fixture
def all_cols():
    from femto.spreadsheet import ColumnData

    default_cols = []
    with open(src_path / 'utils' / 'spreadsheet_columns.txt') as f:
        next(f)
        for line in f:
            tag, name, unit, width, fmt = line.strip().split(', ')
            default_cols.append(ColumnData(tagname=tag, name=name, unit=unit, width=width, format=fmt))
    return default_cols


@pytest.fixture
def list_wg() -> list[Waveguide]:
    PARAM_WG = dict(speed=20, radius=25, pitch=0.080, int_dist=0.007, samplesize=(25, 3))

    coup = [Waveguide(**PARAM_WG) for _ in range(5)]
    for i, wg in enumerate(coup):
        wg.start([-2, i * wg.pitch, 0.035])
        wg.linear([5, 0, 0])
        wg.coupler(dy=(-1) ** i * wg.dy_bend, dz=0, fx=sin)
        wg.coupler(dy=(-1) ** i * wg.dy_bend, dz=0, fx=sin)
        wg.linear([5, 0, 0])
        wg.end()
    return coup


@pytest.fixture
def list_mk() -> list[Marker]:
    PARAM_MK = dict(scan=1, speed=2, speed_pos=5, speed_closed=5, depth=0.000, lx=1, ly=1)
    markers = []
    for x, y in zip(range(4, 8), range(3, 7)):
        m = Marker(**PARAM_MK)
        m.cross([x, y])
        markers.append(m)
    return markers


@pytest.fixture
def ss_param():
    return dict(
        book_name='custom_book_name.xlsx',
        sheet_name='custom_sheet_name',
        columns_names=['name', 'power', 'speed', 'scan', 'depth', 'int_dist', 'yin', 'yout', 'obs'],
    )


@pytest.fixture
def wg_param():
    return dict(
        speed_closed=40,
        radius=40,
        depth=-0.860,
        pitch=0.080,
        samplesize=(25, 25),
    )


@pytest.fixture
def gc_param():
    return dict(
        filename='test_program.pgm',
        laser='PHAROS',
        n_glass=1.4625,
        n_environment=1.000,
        samplesize=(25, 25),
    )


# def device_redd_cols(redd_cols, non_redd_cols, gc_param):
#
#     # redd_cols: the reddundant columns, so all guides have this same attribute
#     dev = Device(**gc_param)
#     ints = iter(range(2, 500))
#     for i_guide in range(10):
#
#         non_rep_value = next(ints)
#
#         start_pt = [-2, 2 + i_guide * 0.08, 0]
#         wg = Waveguide()
#
#         for attribute in redd_cols:
#             setattr(wg, attribute, 50)  # set them all to the same value
#
#         for attribute in non_redd_cols:
#             setattr(wg, attribute, non_rep_value)
#
#         wg.start(start_pt)
#         wg.linear([27, 0, 0])
#         wg.end()
#
#         dev.append(wg)
#
#     return dev


def test_spsh_init_default():
    S = Spreadsheet()
    assert S.columns_names == ['name', 'power', 'speed', 'scan', 'radius', 'int_dist', 'depth', 'yin', 'yout', 'obs']
    assert S.description == ''
    assert S.book_name == 'FABRICATION.xlsx'
    assert S.sheet_name == 'Fabrication'
    assert S.font_name == 'DejaVu Sans Mono'
    assert S.font_size == 11
    assert S.redundant_cols is False
    assert S.new_columns == []
    assert S.metadata == {}

    assert S._workbook.default_format_properties['font_name'] == 'DejaVu Sans Mono'
    assert S._workbook.default_format_properties['font_size'] == 11
    assert S._workbook.filename.name == 'FABRICATION.xlsx'
    assert S._worksheet.name == 'Fabrication'


def test_extra_preamble_info(all_cols, ss_param):
    extra_preamble_info = {
        'wavelength': '1030 nm',
        'laboratory': 'CAPABLE',
        'reprate': '1 MHz',
        'material': 'EAGLE XG',
        'facet': 'bottom',
        'thickness': '900 um',
        'preset': '4',
        'attenuator': '33 %',
        'objective': '1000x',
        'filename': 'My sample 01',
    }

    ss_param['metadata'] = extra_preamble_info
    ss_param['book_name'] = 'extra_preamble_info.xlsx'

    with Spreadsheet(**ss_param) as S:
        S.write([])

    wb = openpyxl.load_workbook(dot_path / 'extra_preamble_info.xlsx')
    worksheet = wb['custom_sheet_name']

    for k, v in extra_preamble_info.items():
        found_extra_par = False
        for row in range(1, 50):
            if worksheet.cell(row=row, column=2).value == k.capitalize().replace('_', ' '):
                found_extra_par = True
                assert worksheet.cell(row=row, column=3).value == v
                break
        assert found_extra_par

    (dot_path / 'extra_preamble_info.xlsx').unlink()


# def test_add_custom_column(ss_param):
#
#     rep_rates = [20, 100, 1000]
#     powers = np.linspace(100, 500, 5)
#     scan_offsets = np.linspace(0.000100, 0.000500, 5)
#
#     dev = Device(filename='test_program.pgm', laser='PHAROS')
#
#     wg_param = dict(
#         speed_closed=40,
#         radius=40,
#         speed=2,
#         scan=10,
#         depth=-0.100,
#         samplesize=(25, 25),
#     )
#
#     for i_guide, (rr, p, so) in enumerate(product(rep_rates, powers, scan_offsets)):
#
#         start_pt = [-2, 2 + i_guide * 0.08, wg_param['depth']]
#
#         wg = Waveguide(**wg_param)
#         wg.power = p  # Can NOT be added inside of the arguments of Waveguide
#         wg.reprate = rr
#         wg.sco = so
#
#         wg.start(start_pt)
#         wg.linear([27, 0, 0])
#         wg.end()
#
#         dev.append(wg)
#
#     ss_param.book_name = 'test_new_column.xlsx'
#     ss_param.columns_names = 'name power reprate sco yin yout obs'
#     ss_param.new_columns = [
#         ('sco', 'Scan offset', 'um', 7, '0.0000'),
#         ('reprate', 'Rep. rate', 'kHz', 7, '0'),
#     ]
#     dev.xlsx(**ss_param)
#
#     wb = openpyxl.load_workbook(dot_path / 'test_new_column.xlsx')
#     worksheet = wb['custom_sheet_name']
#
#     assert worksheet.cell(row=8, column=9).value == 'Scan offset / um'
#     assert worksheet.cell(row=8, column=8).value == 'Rep. rate / kHz'
#
#     for row in range(1, 20):
#         assert worksheet.cell(row=row, column=2).value != 'Rep. rate / kHz'
#         assert worksheet.cell(row=row, column=2).value != 'Rep. rate / MHz'
#
#     (dot_path / 'test_new_column.xlsx').unlink()


def test_write_header(ss_param):
    with Spreadsheet(**ss_param) as S:
        print(S)

    wb = openpyxl.load_workbook(dot_path / 'custom_book_name.xlsx')
    worksheet = wb['custom_sheet_name']

    assert worksheet.cell(row=2, column=6).value == 'Description'
    (dot_path / 'custom_book_name.xlsx').unlink()


def test_write_preamble_default(ss_param):
    del ss_param['columns_names']
    with Spreadsheet(**ss_param) as S:
        print(S)

    wb = openpyxl.load_workbook(dot_path / 'custom_book_name.xlsx')
    worksheet = wb['custom_sheet_name']

    assert worksheet.cell(row=9, column=2).value == 'General'
    assert worksheet.cell(row=10, column=2).value == 'Laboratory'
    assert worksheet.cell(row=11, column=2).value == 'Temperature'
    assert worksheet.cell(row=12, column=2).value == 'Humidity'
    assert worksheet.cell(row=13, column=2).value == 'Date'
    assert worksheet.cell(row=14, column=2).value == 'Start'
    assert worksheet.cell(row=15, column=2).value == 'Filename'

    assert worksheet.cell(row=18, column=2).value == 'Substrate'
    assert worksheet.cell(row=19, column=2).value == 'Material'
    assert worksheet.cell(row=20, column=2).value == 'Facet'
    assert worksheet.cell(row=21, column=2).value == 'Thickness'

    assert worksheet.cell(row=24, column=2).value == 'Laser Parameters'
    assert worksheet.cell(row=25, column=2).value == 'Laser'
    assert worksheet.cell(row=26, column=2).value == 'Wavelength'
    assert worksheet.cell(row=27, column=2).value == 'Duration'
    assert worksheet.cell(row=28, column=2).value == 'Reprate'
    assert worksheet.cell(row=29, column=2).value == 'Attenuator'
    assert worksheet.cell(row=30, column=2).value == 'Preset'

    assert worksheet.cell(row=33, column=2).value == 'Irradiation'
    assert worksheet.cell(row=34, column=2).value == 'Objective'
    assert worksheet.cell(row=35, column=2).value == 'Power'
    assert worksheet.cell(row=36, column=2).value == 'Speed'
    assert worksheet.cell(row=37, column=2).value == 'Scan'
    assert worksheet.cell(row=38, column=2).value == 'Depth'

    for i in range(8, 39):
        assert worksheet.cell(row=i, column=3).value is None
    (dot_path / 'custom_book_name.xlsx').unlink()


def test_add_name_col(ss_param):
    ss_param['columns_names'] = [tag for tag in ss_param['columns_names'] if tag != 'name']
    with Spreadsheet(**ss_param) as S:
        assert S.columns_names[0] == 'name'
    (dot_path / 'custom_book_name.xlsx').unlink()


def test_generate_all_cols_data(all_cols):
    with Spreadsheet() as S:
        assert S.generate_all_cols_data() == all_cols
    (dot_path / 'FABRICATION.xlsx').unlink()


def test_generate_all_cols_with_newcols(all_cols):
    from femto.spreadsheet import ColumnData

    new_cols = [
        ('reprate', 'Rep. rate', 'MHz', '20', '0.000'),
        ('camera', 'CAM', '', '20', 'text'),
        ('cmdrate', 'Command rate', 'pt/s', '25', '0.000'),
    ]
    new_cols_obj = [ColumnData(*nc) for nc in new_cols]

    obs = all_cols.pop(-1)
    all_cols.extend(new_cols_obj[1:])
    all_cols.append(obs)
    with Spreadsheet(new_columns=new_cols) as S:
        assert S.generate_all_cols_data() == all_cols
    (dot_path / 'FABRICATION.xlsx').unlink()


@pytest.mark.parametrize(
    'ncol',
    [
        ('reprate', 'MHz', '20', '0.000'),
        ('camera', 'CAM', '', 'text'),
        ('cmdrate', 'Command rate', 'pt/s', '25', '0.000', 'foo', 'bar'),
    ],
)
def test_new_col_wrong_format(ncol):
    with pytest.raises(ValueError):
        Spreadsheet(new_columns=ncol)


def test_create_structures(list_wg, gc_param, ss_param):
    with Spreadsheet(**ss_param) as spsh:
        spsh.write(list_wg)

    assert (dot_path / 'custom_book_name.xlsx').is_file()
    (dot_path / 'custom_book_name.xlsx').unlink()


@pytest.mark.parametrize('structures', [[], [[]], [[[]]], [[[[]]]]])
def test_write_empty(ss_param, structures):
    with Spreadsheet(**ss_param) as S:
        info, numdata = S._extract_data(structures)
    assert info == []
    assert numdata.size == 0
    (dot_path / 'custom_book_name.xlsx').unlink()


def test_write_ignore_cols(ss_param, list_wg):
    with Spreadsheet(**ss_param, redundant_cols=True) as S:
        info, numdata = S._extract_data(list_wg)

    for row in range(numdata.size):
        for elem in numdata[row]:
            if isinstance(elem, str):
                continue
            else:
                assert elem < 1.1e5

    (dot_path / 'custom_book_name.xlsx').unlink()


def test_write_error(ss_param):
    with pytest.raises(ValueError):
        with Spreadsheet(**ss_param) as S:
            S.write([1, 2, 3])
    (dot_path / 'custom_book_name.xlsx').unlink()


def test_create_formats():
    with Spreadsheet() as S:
        fmt_dict = S._default_formats()

    default_value = {
        'title': {
            'font_color': 'white',
            'bg_color': '#6C5B7B',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'bold': True,
            'text_wrap': True,
        },
        'parname': {
            'bg_color': '#D5CABD',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'bold': True,
            'text_wrap': True,
        },
        'parval': {'text_wrap': True, 'align': 'center', 'valign': 'vcenter', 'border': 1},
        'text': {'align': 'center', 'valign': 'vcenter'},
        'date': {'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': 'DD/MM/YYYY'},
        'time': {'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': 'HH:MM:SS'},
    }
    assert fmt_dict == default_value


def test_add_line_no_format() -> None:
    with Spreadsheet() as S:
        S.add_line(row=1, col=1, data=['ciao'], fmt=None)
        _, fmt = S._worksheet.table[1][1]
        assert fmt.font_size == 11
        assert fmt.font_name == 'DejaVu Sans Mono'
    (dot_path / 'FABRICATION.xlsx').unlink()


@pytest.mark.parametrize(
    'fmt_str, expectation',
    [
        ('title', does_not_raise()),
        (['parname', 'title', 'title'], does_not_raise()),
        (['date', 'date', 'date', 'date'], does_not_raise()),
        (['DATE', 'PARNAME'], does_not_raise()),
        (['null'], pytest.raises(KeyError)),
    ],
)
def test_add_line_keyerror(fmt_str, expectation) -> None:
    S = Spreadsheet()
    data = len(fmt_str) * [1]
    with expectation:
        S.add_line(row=1, col=1, data=data, fmt=fmt_str)


@pytest.mark.parametrize(
    'data, fmt_str, expectation',
    [
        ([1, 2, 3], 'title', does_not_raise()),
        ([1, 2, 3], ['parname', 'title', 'title'], does_not_raise()),
        ([1, '2', 3, 'test'], ['date', 'date', 'date', 'date'], does_not_raise()),
        ([1, 2, 3], ['title', 'parname', 'parname', 'parname'], pytest.raises(ValueError)),
    ],
)
def test_add_line_length_data_fmt(data, fmt_str, expectation) -> None:
    S = Spreadsheet()
    with expectation:
        S.add_line(row=1, col=1, data=data, fmt=fmt_str)


def test_add_line_formula() -> None:
    with Spreadsheet() as S:
        S.add_line(row=1, col=1, data='=SUM(A1:A5)')
        assert S._worksheet.table[1][1].formula
        assert S._worksheet.table[1][1].formula == 'SUM(A1:A5)'
    (dot_path / 'FABRICATION.xlsx').unlink()


def test_add_line_data() -> None:
    with Spreadsheet() as S:
        S.add_line(row=1, col=1, data=[1, 2, 'ciao', 1.2])
        for row in range(5):
            row_dict = S._worksheet.table.get(row, None)
            for col in range(5):
                if row_dict is not None:
                    col_entry = row_dict.get(col, None)
                else:
                    col_entry = None
                print(row, col, col_entry)
        shared_strings = sorted(S._worksheet.str_table.string_table, key=S._worksheet.str_table.string_table.get)

        assert S._worksheet.table[1][1].number
        assert S._worksheet.table[1][1].number == 1
        assert S._worksheet.table[1][2].number
        assert S._worksheet.table[1][2].number == 2
        assert S._worksheet.table[1][3].string
        assert shared_strings[S._worksheet.table[1][3].string] == 'ciao'
        assert S._worksheet.table[1][4].number
        assert S._worksheet.table[1][4].number == 1.2
    (dot_path / 'FABRICATION.xlsx').unlink()
